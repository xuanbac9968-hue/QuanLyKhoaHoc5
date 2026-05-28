"""
Standalone Selenium Test Runner – QuanLyKhoaHoc5  v3
Fixes from v2:
  - TC03: Email unique theo timestamp (tránh trùng lặp DB giữa các lần chạy)
  - TC04[5]: execute_script bypass HTML maxlength attribute trước khi submit
  - TC05: Defensive url_after check (None→''), retry 1 lần, sleep giữa tests
  - TC06: execute_script click (bypass alert overlay che khuất button)
"""

import sys, time, traceback
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ──────────────────────────── CẤU HÌNH ────────────────────────────
BASE_URL     = "http://localhost:5125"
DATA_DIR     = Path(r"D:\QuanLyKhoaHoc5\KiemThu\TestData")
ADMIN_EMAIL  = "admin@nnl.com"
ADMIN_PASS   = "Admin@123"
TIMEOUT      = 12
REPORT_PATH  = Path(r"D:\QuanLyKhoaHoc5\KiemThu\04_katalon_new\report_python.txt")

# ──────────────────────────── HELPERS ─────────────────────────────
def make_driver():
    opts = Options()
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    return webdriver.Chrome(options=opts)

def cell_str(val):
    """Convert Excel cell value to string. Fixes str(0 or '')='' bug."""
    if val is None:
        return ''
    s = str(val)
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace(',', '')

def wait_for(driver, css, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, css))
    )

def wait_page_load(driver, timeout=TIMEOUT):
    """Wait until document.readyState == 'complete'."""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except Exception:
        pass

def elem_visible(driver, css, timeout=3):
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, css))
        )
        return True
    except Exception:
        return False

def has_any_error(driver, timeout=4):
    """Detect both server-side (.alert-danger) and client-side (span.text-danger) errors."""
    for sel in [".alert-danger", ".validation-summary-errors"]:
        try:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            if any(e.is_displayed() and e.text.strip() for e in els):
                return True
        except Exception:
            pass
    try:
        els = driver.find_elements(By.CSS_SELECTOR,
            ".text-danger, span[data-valmsg-for], .field-validation-error")
        if any(e.is_displayed() and e.text.strip() for e in els):
            return True
    except Exception:
        pass
    return False

def login(driver, email, password):
    """Login and wait until redirected away from /Account/Login.
    Raises TimeoutException if login fails (wrong credentials).
    """
    driver.get(BASE_URL + "/Account/Login")
    wait_for(driver, "input#Email").clear()
    driver.find_element(By.CSS_SELECTOR, "input#Email").send_keys(email)
    pw = driver.find_element(By.CSS_SELECTOR, "input#pwInput")
    pw.clear()
    pw.send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    WebDriverWait(driver, TIMEOUT).until(
        lambda d: "/Account/Login" not in d.current_url
    )

def write_result(path, row_idx, col_idx, value):
    wb = load_workbook(path)
    ws = wb.active
    ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
    wb.save(path)
    wb.close()

# ═══════════════════════════════════════════════════════════════════
# TC01 – Đăng nhập
# ═══════════════════════════════════════════════════════════════════
def run_tc01():
    path = DATA_DIR / "TC01_DangNhap.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    results = []

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        stt, mo_ta     = row[0], row[1]
        email          = row[2] or ''
        mat_khau       = cell_str(row[3])
        ghi_nho        = str(row[4] or '').lower()
        ket_qua_mong   = row[5] or ''
        print(f"  TC01 [{stt}] {mo_ta}")

        driver = make_driver()
        ket_qua = "FAIL"
        try:
            driver.get(BASE_URL + "/Account/Login")
            wait_for(driver, "input#Email").send_keys(email)
            driver.find_element(By.CSS_SELECTOR, "input#pwInput").send_keys(mat_khau)
            if ghi_nho == 'true':
                chk = driver.find_element(By.CSS_SELECTOR, "input#GhiNho")
                if not chk.is_selected():
                    chk.click()
            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
            time.sleep(2)
            url = driver.current_url

            if ket_qua_mong == "ThanhCong":
                ket_qua = "PASS" if "/Account/Login" not in url else "FAIL: Vẫn ở trang Login"
            else:
                ok = "/Account/Login" in url or has_any_error(driver)
                ket_qua = "PASS" if ok else "FAIL: Lẽ ra phải ở lại trang Login"
        except Exception as e:
            ket_qua = f"ERROR: {str(e)[:80]}"
        finally:
            try:
                driver.quit()
            except Exception:
                pass

        write_result(path, i + 1, 6, ket_qua)
        results.append((stt, mo_ta, ket_qua_mong, ket_qua))
        print(f"    → {ket_qua}")
    return results

# ═══════════════════════════════════════════════════════════════════
# TC02 – Đổi mật khẩu
# ═══════════════════════════════════════════════════════════════════
def run_tc02():
    path = DATA_DIR / "TC02_DoiMatKhau.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    results = []

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        stt, mo_ta    = row[0], row[1]
        email_login   = row[2] or ''
        mk_login      = cell_str(row[3])
        mk_cu         = cell_str(row[4])
        mk_moi        = cell_str(row[5])
        xac_nhan      = cell_str(row[6])
        ket_qua_mong  = row[7] or ''
        print(f"  TC02 [{stt}] {mo_ta}")

        driver = make_driver()
        ket_qua = "FAIL"
        try:
            try:
                login(driver, email_login, mk_login)
            except TimeoutException:
                if ket_qua_mong == "ThatBai":
                    ket_qua = "PASS (login thất bại như mong đợi)"
                else:
                    ket_qua = "ERROR: Không thể đăng nhập"
                write_result(path, i + 1, 8, ket_qua)
                results.append((stt, mo_ta, ket_qua_mong, ket_qua))
                print(f"    → {ket_qua}")
                continue

            driver.get(BASE_URL + "/Account/ChangePassword")
            wait_for(driver, "input#MatKhauCu").send_keys(mk_cu)
            driver.find_element(By.CSS_SELECTOR, "input#MatKhauMoi").send_keys(mk_moi)
            driver.find_element(By.CSS_SELECTOR, "input#XacNhanMatKhau").send_keys(xac_nhan)
            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
            time.sleep(2.5)

            has_success = elem_visible(driver, ".alert.alert-success", timeout=4)
            has_error   = has_any_error(driver)

            if ket_qua_mong == "ThanhCong":
                ket_qua = "PASS" if has_success else "FAIL: Không thấy alert thành công"
            else:
                ket_qua = "PASS" if has_error else "FAIL: Lẽ ra phải có lỗi"

        except Exception as e:
            ket_qua = f"ERROR: {str(e)[:80]}"
        finally:
            try:
                driver.quit()
            except Exception:
                pass

        write_result(path, i + 1, 8, ket_qua)
        results.append((stt, mo_ta, ket_qua_mong, ket_qua))
        print(f"    → {ket_qua}")
    return results

# ═══════════════════════════════════════════════════════════════════
# TC03 – Tạo tài khoản (Admin modal)
# FIX v3: Email unique theo timestamp (tránh trùng lặp DB)
# ═══════════════════════════════════════════════════════════════════
def run_tc03():
    path = DATA_DIR / "TC03_TaoTaiKhoan.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    results = []

    # FIX v3: Timestamp suffix để email không bị trùng DB giữa các lần chạy
    ts = datetime.now().strftime("%H%M%S")

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        stt, mo_ta    = row[0], row[1]
        ho_ten        = row[2] or ''
        email         = row[3] or ''
        vai_tro       = row[4] or ''
        mat_khau      = cell_str(row[5])
        xac_nhan      = cell_str(row[6])
        ket_qua_mong  = row[7] or ''

        # FIX v3: ThanhCong cases → append timestamp to avoid duplicate email
        if ket_qua_mong == "ThanhCong" and email and '@' in email:
            local, domain = email.split('@', 1)
            email = f"{local}_{ts}@{domain}"

        print(f"  TC03 [{stt}] {mo_ta}")

        driver = make_driver()
        ket_qua = "FAIL"
        try:
            login(driver, ADMIN_EMAIL, ADMIN_PASS)
            driver.get(BASE_URL + "/TaiKhoan")
            WebDriverWait(driver, TIMEOUT).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[onclick='openCreateModal()']"))
            ).click()
            time.sleep(0.8)

            wait_for(driver, "input#cr-hoTen").clear()
            driver.find_element(By.CSS_SELECTOR, "input#cr-hoTen").send_keys(ho_ten)
            driver.find_element(By.CSS_SELECTOR, "input#cr-email").clear()
            driver.find_element(By.CSS_SELECTOR, "input#cr-email").send_keys(email)
            Select(driver.find_element(By.CSS_SELECTOR, "select#cr-vaiTro")).select_by_value(vai_tro)
            driver.find_element(By.CSS_SELECTOR, "input#cr-matKhau").send_keys(mat_khau)
            driver.find_element(By.CSS_SELECTOR, "input#cr-xacNhan").send_keys(xac_nhan)
            driver.find_element(By.CSS_SELECTOR, "button#btn-create-submit").click()
            time.sleep(2.5)

            has_toast = elem_visible(driver, "#toast-container .toast.bg-success", timeout=4)
            alert_els = driver.find_elements(By.CSS_SELECTOR, "div#create-alert")
            has_err_msg = (bool(alert_els) and
                           "d-none" not in (alert_els[0].get_attribute("class") or "") and
                           alert_els[0].text.strip() != "")

            if ket_qua_mong == "ThanhCong":
                ket_qua = "PASS" if has_toast else (
                    "FAIL: Email tồn tại hoặc lỗi khác" if has_err_msg
                    else "FAIL: Không thấy toast success")
            else:
                ket_qua = "PASS" if has_err_msg else "FAIL: Lẽ ra phải có lỗi"

        except Exception as e:
            ket_qua = f"ERROR: {str(e)[:80]}"
        finally:
            try:
                driver.quit()
            except Exception:
                pass

        write_result(path, i + 1, 8, ket_qua)
        results.append((stt, mo_ta, ket_qua_mong, ket_qua))
        print(f"    → {ket_qua}")
    return results

# ═══════════════════════════════════════════════════════════════════
# TC04 – Thêm khóa học
# FIX v3: bypass HTML maxlength attribute cho TenKhoaHoc > 200 ký tự
# ═══════════════════════════════════════════════════════════════════
def run_tc04():
    path = DATA_DIR / "TC04_ThemKhoaHoc.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    results = []

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        stt, mo_ta    = row[0], row[1]
        ten_kh        = row[2] or ''
        mo_ta_kh      = row[3] or ''
        ngon_ngu      = row[4] or ''
        trinh_do      = row[5] or ''
        hoc_phi       = cell_str(row[6])
        thoi_luong    = cell_str(row[7])
        so_buoi       = cell_str(row[8])
        thoi_gian     = cell_str(row[9])
        trang_thai    = row[10] or ''
        ket_qua_mong  = row[11] or ''
        print(f"  TC04 [{stt}] {mo_ta}")

        driver = make_driver()
        ket_qua = "FAIL"
        try:
            login(driver, ADMIN_EMAIL, ADMIN_PASS)
            driver.get(BASE_URL + "/KhoaHoc/Create")

            # FIX v3: bypass HTML maxlength khi TenKhoaHoc > 200 ký tự
            ten_kh_el = wait_for(driver, "input#TenKhoaHoc")
            if len(str(ten_kh)) > 200:
                # Xóa maxlength attribute và set value trực tiếp qua JS
                driver.execute_script(
                    "arguments[0].removeAttribute('maxlength'); "
                    "arguments[0].value = arguments[1];",
                    ten_kh_el, str(ten_kh)
                )
                # Trigger input event để jQuery Unobtrusive Validation nhận ra giá trị
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('input', {bubbles:true})); "
                    "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));",
                    ten_kh_el
                )
            else:
                ten_kh_el.send_keys(ten_kh)

            driver.find_element(By.CSS_SELECTOR, "textarea#MoTa").send_keys(mo_ta_kh)

            try:
                Select(driver.find_element(By.CSS_SELECTOR, "select#NgonNgu")).select_by_visible_text(ngon_ngu)
            except Exception:
                pass

            try:
                Select(driver.find_element(By.CSS_SELECTOR, "select#TrinhDo")).select_by_value(trinh_do)
            except Exception:
                pass

            hp_el = driver.find_element(By.CSS_SELECTOR, "input#HocPhi")
            driver.execute_script("arguments[0].value = '';", hp_el)
            if hoc_phi:
                hp_el.send_keys(hoc_phi)

            tl_el = driver.find_element(By.CSS_SELECTOR, "input#ThoiLuong")
            driver.execute_script("arguments[0].value = '';", tl_el)
            if thoi_luong:
                tl_el.send_keys(thoi_luong)

            for css, val in [("input#SoBuoiMoiTuan", so_buoi), ("input#ThoiGianMoiBuoi", thoi_gian)]:
                el = driver.find_element(By.CSS_SELECTOR, css)
                driver.execute_script("arguments[0].value = '';", el)
                if val:
                    el.send_keys(val)

            try:
                Select(driver.find_element(By.CSS_SELECTOR, "select#TrangThai")).select_by_value(trang_thai)
            except Exception:
                pass

            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
            time.sleep(2.5)

            url     = driver.current_url
            has_err = has_any_error(driver)

            if ket_qua_mong == "ThanhCong":
                if "/KhoaHoc/Create" not in url and not has_err:
                    ket_qua = "PASS"
                else:
                    err_text = ""
                    try:
                        els = driver.find_elements(By.CSS_SELECTOR, ".alert-danger, .text-danger")
                        err_text = next((e.text.strip()[:60] for e in els if e.text.strip()), "")
                    except Exception:
                        pass
                    ket_qua = f"FAIL: {err_text or 'Vẫn ở trang Create'}"
            else:
                still_on_create = "/KhoaHoc/Create" in url
                ket_qua = "PASS" if (has_err or still_on_create) else "FAIL: Lẽ ra phải có lỗi validation"

        except Exception as e:
            ket_qua = f"ERROR: {str(e)[:80]}"
        finally:
            try:
                driver.quit()
            except Exception:
                pass

        write_result(path, i + 1, 12, ket_qua)
        results.append((stt, mo_ta, ket_qua_mong, ket_qua))
        print(f"    → {ket_qua}")
    return results

# ═══════════════════════════════════════════════════════════════════
# TC05 – Tạo thanh toán (HocVien)
# FIX v3: Defensive url_after=None, retry 1 lần, sleep giữa tests
# ═══════════════════════════════════════════════════════════════════
def _run_tc05_item(email_hv, mk_hv, kh_id, phuong_thuc, ghi_chu, ket_qua_mong):
    """Run one TC05 sub-test with its own driver. Returns ket_qua string."""
    driver = make_driver()
    try:
        login(driver, email_hv, mk_hv)

        driver.get(BASE_URL + f"/ThanhToan/TaoYeuCau?khoaHocId={kh_id}")
        wait_page_load(driver)
        time.sleep(0.8)

        # FIX v3: Defensive None check — driver.current_url có thể trả về None nếu browser crash
        url_after = driver.current_url or ''

        if "TaoYeuCau" not in url_after:
            if ket_qua_mong == "ThatBai":
                return "PASS"
            else:
                if "/ThanhToan/CuaToi" in url_after:
                    has_warning = elem_visible(driver, ".alert.alert-warning", timeout=3)
                    rows_table = []
                    try:
                        rows_table = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                    except Exception:
                        pass
                    if has_warning or len(rows_table) > 0:
                        return "PASS (thanh toán đã tồn tại — ChoPheduyet)"
                return f"FAIL: Redirect tới {url_after[:60]}"
        else:
            # Trang TaoYeuCau load thành công — chọn phương thức và submit
            if phuong_thuc == 'TienMat':
                radio = driver.find_element(By.CSS_SELECTOR, "input#ptTienMat")
            else:
                radio = driver.find_element(By.CSS_SELECTOR, "input#ptChuyenKhoan")
            if not radio.is_selected():
                driver.execute_script("arguments[0].click();", radio)

            ghi_chu_el = driver.find_element(By.CSS_SELECTOR, "textarea#GhiChu")
            ghi_chu_el.clear()
            ghi_chu_el.send_keys(str(ghi_chu))

            driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
            wait_page_load(driver)
            time.sleep(0.8)

            url2        = driver.current_url or ''
            has_success = (elem_visible(driver, ".alert.alert-success", timeout=3) or
                           "TaoYeuCau" not in url2)

            if ket_qua_mong == "ThanhCong":
                return "PASS" if has_success else "FAIL: Không thấy success"
            else:
                return "PASS" if not has_success else "FAIL: Lẽ ra phải thất bại"
    finally:
        try:
            driver.quit()
        except Exception:
            pass


def run_tc05():
    path = DATA_DIR / "TC05_TaoThanhToan.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    results = []

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        stt, mo_ta    = row[0], row[1]
        email_hv      = row[2] or ''
        mk_hv         = cell_str(row[3])
        kh_id         = cell_str(row[4])
        phuong_thuc   = row[5] or 'TienMat'
        ghi_chu       = row[6] or ''
        ket_qua_mong  = row[7] or ''
        print(f"  TC05 [{stt}] {mo_ta}")

        ket_qua = "FAIL"
        # FIX v3: Retry 1 lần nếu gặp exception (timeout / crash)
        for attempt in range(2):
            try:
                ket_qua = _run_tc05_item(
                    email_hv, mk_hv, kh_id, phuong_thuc, ghi_chu, ket_qua_mong
                )
                break  # thành công → thoát retry loop
            except Exception as e:
                ket_qua = f"ERROR: {str(e)[:80]}"
                if attempt == 0:
                    print(f"    ⟳ Retry lần 2 (sau 5s)...")
                    time.sleep(5)

        write_result(path, i + 1, 8, ket_qua)
        results.append((stt, mo_ta, ket_qua_mong, ket_qua))
        print(f"    → {ket_qua}")
        # FIX v3: Pause nhỏ giữa các TC05 test để app không bị quá tải
        time.sleep(2)

    return results

# ═══════════════════════════════════════════════════════════════════
# TC06 – Phân công giảng viên (Admin)
# FIX v3: execute_script click button (bypass alert overlay)
# ═══════════════════════════════════════════════════════════════════
def run_tc06():
    path = DATA_DIR / "TC06_PhanCong.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    results = []

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        stt, mo_ta    = row[0], row[1]
        kh_id         = cell_str(row[2])
        gv_id         = cell_str(row[3])
        ket_qua_mong  = row[5] or ''
        if not kh_id:
            continue
        print(f"  TC06 [{stt}] {mo_ta}")

        driver = make_driver()
        ket_qua = "FAIL"
        try:
            login(driver, ADMIN_EMAIL, ADMIN_PASS)
            driver.get(BASE_URL + "/Admin/PhanCong")
            time.sleep(1.5)

            xpath_select = (f"//form[.//input[@name='KhoaHocId' and @value='{kh_id}']]"
                            f"//select[@name='GiangVienId']")
            sel_el = WebDriverWait(driver, TIMEOUT).until(
                EC.visibility_of_element_located((By.XPATH, xpath_select))
            )
            Select(sel_el).select_by_value(gv_id)

            xpath_btn = (f"//form[.//input[@name='KhoaHocId' and @value='{kh_id}']]"
                         f"//button[@type='submit']")
            btn = driver.find_element(By.XPATH, xpath_btn)

            # FIX v3: scroll vào view rồi click bằng JS để bypass bất kỳ overlay nào
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(2)

            has_success = elem_visible(driver, ".alert.alert-success", timeout=4)
            ket_qua = "PASS" if has_success else "FAIL: Không thấy alert thành công"

        except Exception as e:
            ket_qua = f"ERROR: {str(e)[:80]}"
        finally:
            try:
                driver.quit()
            except Exception:
                pass

        write_result(path, i + 1, 6, ket_qua)
        results.append((stt, mo_ta, ket_qua_mong, ket_qua))
        print(f"    → {ket_qua}")
    return results

# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    start = datetime.now()
    all_results = {}

    tcs = [
        ("TC01 – Đăng nhập",      run_tc01),
        ("TC02 – Đổi mật khẩu",   run_tc02),
        ("TC03 – Tạo tài khoản",   run_tc03),
        ("TC04 – Thêm khóa học",   run_tc04),
        ("TC05 – Tạo thanh toán",  run_tc05),
        ("TC06 – Phân công GV",    run_tc06),
    ]

    for idx, (name, fn) in enumerate(tcs):
        print(f"\n{'═'*55}")
        print(f"  {name}")
        print('═' * 55)

        # Pause trước TC05 để app recover sau khi TC01–TC04 chạy xong
        if idx == 4:
            print("  ⏳ Chờ 3s trước TC05 để app recover...")
            time.sleep(3)

        try:
            all_results[name] = fn()
        except Exception as e:
            print(f"  [FATAL] {e}")
            all_results[name] = []

    elapsed = (datetime.now() - start).total_seconds()
    lines = []
    lines.append("=" * 60)
    lines.append("  BÁO CÁO KIỂM THỬ v3 – QuanLyKhoaHoc5")
    lines.append(f"  Thời gian: {start.strftime('%d/%m/%Y %H:%M:%S')}")
    lines.append(f"  Tổng thời gian: {elapsed:.0f}s")
    lines.append("=" * 60)

    grand_pass = grand_fail = grand_error = 0

    for tc_name, tc_rows in all_results.items():
        tp = sum(1 for r in tc_rows if str(r[3]).startswith("PASS"))
        tf = sum(1 for r in tc_rows if str(r[3]).startswith("FAIL"))
        te = sum(1 for r in tc_rows if str(r[3]).startswith("ERROR"))
        grand_pass += tp; grand_fail += tf; grand_error += te

        lines.append(f"\n▶ {tc_name}")
        lines.append(f"  Tổng: {len(tc_rows)} | ✅ PASS: {tp} | ❌ FAIL: {tf} | ⚠ ERROR: {te}")
        for r in tc_rows:
            icon = "✅" if str(r[3]).startswith("PASS") else ("⚠" if str(r[3]).startswith("ERROR") else "❌")
            lines.append(f"  {icon} [{r[0]}] {r[1]}")
            if not str(r[3]).startswith("PASS"):
                lines.append(f"      → {r[3]}")

    total = grand_pass + grand_fail + grand_error
    pct   = int(grand_pass / total * 100) if total else 0
    lines.append("\n" + "=" * 60)
    lines.append(f"  TỔNG KẾT: {total} test cases")
    lines.append(f"  ✅ PASSED : {grand_pass}")
    lines.append(f"  ❌ FAILED : {grand_fail}")
    lines.append(f"  ⚠  ERRORS : {grand_error}")
    lines.append(f"  Tỉ lệ pass: {pct}%")
    lines.append("=" * 60)
    lines.append(f"  Kết quả chi tiết → {DATA_DIR}")
    lines.append("=" * 60)

    report = "\n".join(lines)
    print("\n\n" + report)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"\n  📄 Báo cáo: {REPORT_PATH}")

if __name__ == "__main__":
    main()
