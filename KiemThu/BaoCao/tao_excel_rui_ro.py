"""
tao_excel_rui_ro.py
Tạo file Excel đầy đủ: Phân tích rủi ro + Chiến lược kiểm thử + Bảng bug Jira
Chạy: python tao_excel_rui_ro.py
Output: phan_tich_rui_ro.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabel
import os

OUTPUT = os.path.join(os.path.dirname(__file__), "phan_tich_rui_ro.xlsx")

# ── Color palette ─────────────────────────────────────────────────────────
C = {
    "header_blue":    PatternFill("solid", fgColor="1565C0"),
    "header_dark":    PatternFill("solid", fgColor="263238"),
    "header_green":   PatternFill("solid", fgColor="2E7D32"),
    "header_orange":  PatternFill("solid", fgColor="E65100"),
    "header_purple":  PatternFill("solid", fgColor="4A148C"),
    "row_light":      PatternFill("solid", fgColor="F8F9FA"),
    "row_alt":        PatternFill("solid", fgColor="FFFFFF"),
    "risk_critical":  PatternFill("solid", fgColor="FFCDD2"),
    "risk_high":      PatternFill("solid", fgColor="FFE0B2"),
    "risk_medium":    PatternFill("solid", fgColor="FFF9C4"),
    "risk_low":       PatternFill("solid", fgColor="E8F5E9"),
    "pass_green":     PatternFill("solid", fgColor="C8E6C9"),
    "fail_red":       PatternFill("solid", fgColor="FFCDD2"),
    "warn_yellow":    PatternFill("solid", fgColor="FFF3E0"),
    "blue_row":       PatternFill("solid", fgColor="E3F2FD"),
    "title_bg":       PatternFill("solid", fgColor="0D47A1"),
}

WHITE   = Font(color="FFFFFF", bold=True, size=11)
BOLD    = Font(bold=True, size=11)
BOLD10  = Font(bold=True, size=10)
NORMAL  = Font(size=10)
SMALL   = Font(size=9)
TITLE_F = Font(color="FFFFFF", bold=True, size=14)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(border_style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)

def med_border():
    s = Side(border_style="medium", color="546E7A")
    return Border(left=s, right=s, top=s, bottom=s)

def set_row(ws, row, values, fills=None, fonts=None, aligns=None, height=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin_border()
        if fills:
            f = fills if not isinstance(fills, list) else (fills[col-1] if col-1 < len(fills) else None)
            if f: c.fill = f
        if fonts:
            fn = fonts if not isinstance(fonts, list) else (fonts[col-1] if col-1 < len(fonts) else NORMAL)
            c.font = fn if fn else NORMAL
        if aligns:
            al = aligns if not isinstance(aligns, list) else (aligns[col-1] if col-1 < len(aligns) else LEFT)
            c.alignment = al if al else LEFT
        else:
            c.alignment = LEFT
    if height:
        ws.row_dimensions[row].height = height

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, text, cols, fill=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill or C["title_bg"]
    c.font = TITLE_F
    c.alignment = CENTER
    c.border = med_border()
    ws.row_dimensions[row].height = 36

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
#  SHEET 1: PHÂN TÍCH RỦI RO
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Phân Tích Rủi Ro"

# Row 1: Title
title_row(ws1, 1, "📊 BẢNG PHÂN TÍCH RỦI RO – QUANLYKHOAHOC5 (NNL Language Center)", 10)

# Row 2: Subtitle
ws1.merge_cells("A2:J2")
c = ws1["A2"]
c.value = "Ngày: 2026-05-28  |  Phiên bản: 1.0  |  Người lập: Nguyễn Xuân Bắc  |  Công cụ: SonarQube, ZAP, Katalon, JMeter"
c.fill = PatternFill("solid", fgColor="1976D2")
c.font = Font(color="FFFFFF", size=10, italic=True)
c.alignment = CENTER
ws1.row_dimensions[2].height = 22

# Row 3: blank
ws1.row_dimensions[3].height = 8

# Row 4: Headers
headers = [
    "ID", "Chức năng\nkiểm thử", "Mô tả rủi ro",
    "Xác suất\n(1-5)", "Mức độ ảnh hưởng\n(1-5)", "Mức Rủi Ro\n= XS × MĐAH",
    "Mức độ\nưu tiên", "Phương pháp\nkiểm thử", "Kết quả\nkiểm thử",
    "Trạng thái\nxử lý"
]
set_row(ws1, 4, headers,
        fills=C["header_blue"],
        fonts=WHITE,
        aligns=CENTER,
        height=44)

col_widths(ws1, [6, 18, 38, 12, 14, 13, 13, 22, 16, 14])

# Risk data
RISKS = [
    # (ID, Module, Mô tả rủi ro, XS, MD, Strategy, KQ, Status)
    ("R-01","Đăng Nhập / Xác thực","SQL Injection qua form đăng nhập – kẻ tấn công bypass auth",4,5,"SQL Injection – CRITICAL","Kiểm thử penetration (Katalon + ZAP)","TC01 PASSED\nZAP: 0 High","Đang xử lý"),
    ("R-02","Đăng Nhập / Xác thực","Brute-force mật khẩu – không có rate limiting / CAPTCHA",3,5,"Brute Force – CRITICAL","Kiểm tra xác thực nhiều lần, đo rate limiting","TC01 PASSED\nRate: chưa có","Mở"),
    ("R-03","Quản Lý Khóa Học","Mất dữ liệu khóa học khi xóa cascade không đúng logic",3,4,"Data Integrity – HIGH","Unit test, Integration test, DB integrity check","531 tests PASSED","Hoàn thành"),
    ("R-04","Quản Lý Khóa Học","Phân quyền sai: HocVien truy cập trang Admin quản lý",4,5,"Authorization – CRITICAL","Kiểm tra từng role, test privilege escalation","TC04 PASSED\nRoles: OK","Hoàn thành"),
    ("R-05","Đăng Ký Học Viên","Đăng ký trùng lặp cho cùng khóa học cùng HocVien",3,3,"Duplicate Entry – MEDIUM","Unit test: DangKyService, unique constraint DB","531 tests PASSED","Hoàn thành"),
    ("R-06","Đăng Ký Học Viên","Không kiểm soát slot – đăng ký vượt số chỗ tối đa",2,4,"Over-enrollment – HIGH","Test boundary: đăng ký vượt SoLuongHocVienToiDa","Unit test PASSED","Hoàn thành"),
    ("R-07","Lịch Học","Xung đột lịch: 2 khóa học cùng phòng cùng giờ",3,4,"Schedule Conflict – HIGH","Kiểm tra conflict detection trong PhanCong","TC06 PASSED\nConflict: OK","Hoàn thành"),
    ("R-08","Lịch Học","ExpandToSessions() sai khi NgayKetThuc < NgayBatDau",2,3,"Data Validation – MEDIUM","Unit test: LichHocHelper edge cases","Coverage80 test","Mở (NNL-010)"),
    ("R-09","Thanh Toán","CSRF attack: POST /ThanhToan/Duyet không có AntiForgeryToken",4,5,"CSRF – CRITICAL","ZAP active scan, manual CSRF test","ZAP: 0 High\nCSRF: chưa fix","Mở (NNL-004)"),
    ("R-10","Thanh Toán","File upload độc hại: rename .exe → .jpg để upload malware",3,5,"File Upload – CRITICAL","Test với file giả mạo, kiểm tra MIME type","TC05 PASSED\nMIME: chưa fix","Mở (NNL-011)"),
    ("R-11","Thanh Toán","Race condition: 2 admin duyệt cùng lúc → double approve",2,4,"Concurrency – HIGH","Concurrent test với JMeter (200 users)","200u: 0 errors","Mở (NNL-003)"),
    ("R-12","Chat AI / Gợi Ý","API Grok bị lộ token nếu commit lên public repository",3,4,"Secret Exposure – HIGH","Kiểm tra .gitignore, git history, secret scan","Token ẩn trong\nappsettings","Đã xử lý (NNL-007)"),
    ("R-13","Chat AI / Gợi Ý","NullRef khi cookie xác thực hết hạn giữa request chat",3,4,"Null Reference – HIGH","Unit test ChatService null cases","Coverage80 test","Mở (NNL-001)"),
    ("R-14","Chat AI / Gợi Ý","HttpClient resource leak trong ChatService – socket exhaustion",3,4,"Resource Leak – HIGH","Load test 300 users, monitor connections","300u: 0 errors","Mở (NNL-006)"),
    ("R-15","Hiệu năng / Tải","P95 response > 1000ms khi 200 users – SLA có thể bị vi phạm",3,3,"Performance – MEDIUM","JMeter: 100/200/300 concurrent users","P95 200u: 1027ms\n(SLA: 2000ms OK)","Theo dõi"),
]

risk_level_map = {
    range(1,9):  ("THẤP",    C["risk_low"],  "🟢"),
    range(9,13): ("TRUNG BÌNH", C["risk_medium"], "🟡"),
    range(13,17):("CAO",     C["risk_high"], "🟠"),
    range(17,26):("NGHIÊM TRỌNG", C["risk_critical"], "🔴"),
}

def get_risk(xs, md):
    score = xs * md
    for r, (lbl, fill, icon) in risk_level_map.items():
        if score in r:
            return score, lbl, fill, icon
    return score, "NGHIÊM TRỌNG", C["risk_critical"], "🔴"

status_fill = {"Hoàn thành":C["pass_green"],"Đang xử lý":C["warn_yellow"],"Mở":C["fail_red"],"Theo dõi":C["blue_row"],"Đã xử lý":C["pass_green"]}

for i, (rid, mod, desc, xs, md, sev_label, strategy, kq, status) in enumerate(RISKS):
    row = 5 + i
    score, lbl, rfill, icon = get_risk(xs, md)
    base_fill = C["row_light"] if i % 2 == 0 else C["row_alt"]
    sfill = status_fill.get(status, C["row_alt"])

    vals = [rid, mod, desc, xs, md, score, f"{icon} {lbl}", strategy, kq, status]
    fills = [base_fill, base_fill, base_fill, base_fill, base_fill, rfill, rfill, base_fill, base_fill, sfill]
    fonts = [BOLD10, BOLD10, NORMAL, BOLD, BOLD, BOLD, BOLD10, SMALL, SMALL, BOLD10]
    aligns = [CENTER, CENTER, LEFT, CENTER, CENTER, CENTER, CENTER, LEFT, LEFT, CENTER]

    set_row(ws1, row, vals, fills=fills, fonts=fonts, aligns=aligns, height=48)

ws1.row_dimensions[5 + len(RISKS)].height = 8

# Summary table
sr = 6 + len(RISKS)
ws1.merge_cells(f"A{sr}:J{sr}")
c = ws1.cell(row=sr, column=1, value="📋 TỔNG HỢP MỨC ĐỘ RỦI RO")
c.fill = C["header_dark"]; c.font = TITLE_F; c.alignment = CENTER
ws1.row_dimensions[sr].height = 30

sr += 1
summary_data = [
    ("🔴 Nghiêm trọng (17-25)", len([r for r in RISKS if r[3]*r[4]>=17]), "Xử lý ngay – sprint hiện tại"),
    ("🟠 Cao (13-16)",          len([r for r in RISKS if 13<=r[3]*r[4]<17]), "Xử lý trong sprint này"),
    ("🟡 Trung bình (9-12)",    len([r for r in RISKS if 9<=r[3]*r[4]<13]),  "Lên kế hoạch xử lý"),
    ("🟢 Thấp (1-8)",           len([r for r in RISKS if r[3]*r[4]<9]),      "Theo dõi, thấp ưu tiên"),
]
set_row(ws1, sr, ["Mức rủi ro", "Số lượng", "Hành động đề xuất"], C["header_blue"], WHITE, CENTER, 26)
for j, (lbl, cnt, action) in enumerate(summary_data):
    set_row(ws1, sr+1+j, [lbl, cnt, action],
            fills=[C["row_light"] if j%2==0 else C["row_alt"]]*3,
            aligns=[LEFT, CENTER, LEFT], height=22)

ws1.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
#  SHEET 2: CHIẾN LƯỢC KIỂM THỬ DỰA TRÊN RỦI RO
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Chiến Lược Kiểm Thử")

title_row(ws2, 1, "🎯 CHIẾN LƯỢC KIỂM THỬ DỰA TRÊN RỦI RO – QuanLyKhoaHoc5", 9, C["header_green"])

ws2.merge_cells("A2:I2")
c = ws2["A2"]
c.value = "Risk-Based Testing (RBT) – Tập trung kiểm thử vào các chức năng có rủi ro cao nhất"
c.fill = PatternFill("solid", fgColor="388E3C"); c.font = Font(color="FFFFFF", size=10, italic=True); c.alignment = CENTER

col_widths(ws2, [6, 18, 12, 14, 18, 16, 16, 16, 18])

ws2.row_dimensions[3].height = 8

headers2 = ["ID", "Chức năng", "Rủi ro\nưu tiên", "Kỹ thuật\nkiểm thử", "Công cụ", "Test Cases\nliên quan", "Tổng\ntest", "Pass/\nFail", "Ghi chú"]
set_row(ws2, 4, headers2, C["header_green"], WHITE, CENTER, 40)

STRATEGIES = [
    ("S-01","Đăng Nhập / Auth","CRITICAL",
     "Black-box, Boundary Value\nEquivalence Partition\nPenetration Testing",
     "Katalon Studio\nOWASP ZAP",
     "TC01_DangNhap","TC: 4 sub-cases","4/4 PASS","Priority 1: kiểm thử đầu tiên"),
    ("S-02","Quản Lý Khóa Học","HIGH",
     "CRUD testing\nRole-based access\nBoundary: pagination",
     "Katalon Studio\nxUnit (531 tests)",
     "TC04_ThemKhoaHoc","Tests: 45","45/45 PASS","CRUD đủ 4 operations"),
    ("S-03","Đăng Ký Học Viên","MEDIUM",
     "Equivalence Partition\nDuplicate detection\nBoundary: max slots",
     "xUnit + Moq\nKatalon Studio",
     "TC03_TaoTaiKhoan","Tests: 38","38/38 PASS","Kiểm tra constraint DB"),
    ("S-04","Lịch Học","HIGH",
     "Conflict detection\nDate range validation\nEdge cases: Sun/holiday",
     "xUnit: LichHocHelper\nKatalon Studio",
     "TC04 (lịch)\nUnit: LichHocHelper","Tests: 28","28/28 PASS","Còn lỗi NNL-010, NNL-020"),
    ("S-05","Thanh Toán","CRITICAL",
     "File upload security\nCSRF testing\nConcurrency: 200 users",
     "ZAP\nJMeter 200u\nKatalon Studio",
     "TC05_TaoThanhToan","TC: 3 sub-cases\nJMeter: 2000 req","PASS\n0 errors","Còn NNL-004, NNL-011"),
    ("S-06","Chat AI / Gợi Ý","HIGH",
     "API error handling\nNull safety testing\nLoad: 300 users",
     "xUnit: ChatService\nGoiYService\nJMeter 300u",
     "Unit: Coverage80Tests","Tests: 75","75/75 PASS","Token management: NNL-007 done"),
    ("S-07","Hiệu năng toàn hệ thống","MEDIUM",
     "Load testing\nStress testing\nPerformance profiling",
     "JMeter 5.6.3\n100/200/300 users",
     "Tất cả endpoints","6000 requests\n(3 scenarios)","0 errors\nP95 < 2000ms","SLA đạt được"),
    ("S-08","Bảo mật toàn diện","HIGH",
     "DAST scanning\nSecurity headers check\nVulnerability assessment",
     "OWASP ZAP 2.16.0\nSonarQube",
     "Toàn bộ ứng dụng","0 High alerts\n3 Medium\n1 Low","PASS (0 High)","Medium cần fix: CSP, SRI"),
]

fills_s = [C["risk_critical"], C["risk_high"], C["risk_medium"], C["risk_high"],
           C["risk_critical"], C["risk_high"], C["risk_medium"], C["risk_high"]]

for i, (sid, fn, pri, tech, tools, tc, cnt, result, note) in enumerate(STRATEGIES):
    row = 5 + i
    base = C["row_light"] if i%2==0 else C["row_alt"]
    rfill = fills_s[i]
    pfill = C["pass_green"] if "PASS" in result else C["fail_red"]

    set_row(ws2, row,
            [sid, fn, pri, tech, tools, tc, cnt, result, note],
            fills=[base, base, rfill, base, base, base, base, pfill, base],
            fonts=[BOLD10, BOLD10, BOLD10, SMALL, SMALL, SMALL, SMALL, BOLD10, SMALL],
            aligns=[CENTER, CENTER, CENTER, LEFT, LEFT, LEFT, CENTER, CENTER, LEFT],
            height=58)

ws2.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
#  SHEET 3: BẢNG BUG JIRA
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Bug Tracker (Jira)")

title_row(ws3, 1, "🐛 BẢNG THEO DÕI BUG – JIRA | xuanbac9968.atlassian.net | Project: NNL", 10, C["header_orange"])

ws3.merge_cells("A2:J2")
c = ws3["A2"]
c.value = "Import từ SonarQube 26.5.0 + OWASP ZAP | Tổng: 20 issues (10 Bugs + 10 Vulnerabilities) | 2026-05-28"
c.fill = PatternFill("solid", fgColor="BF360C"); c.font = Font(color="FFFFFF", size=10, italic=True); c.alignment = CENTER

col_widths(ws3, [8, 10, 10, 12, 38, 20, 8, 10, 14, 18])
ws3.row_dimensions[3].height = 8

headers3 = ["Jira ID", "Loại", "Severity", "Ưu tiên", "Mô tả vấn đề", "Component:Line", "Effort", "Status", "Assignee", "Nguồn phát hiện"]
set_row(ws3, 4, headers3, C["header_orange"], WHITE, CENTER, 40)

BUGS = [
    ("NNL-001","Bug","CRITICAL","High","NullReferenceException khi cookie xác thực hết hạn giữa request API chat","ChatController.cs:47","30m","Open","xuanbac9968","SonarQube"),
    ("NNL-002","Bug","MAJOR","High","Potential SQL Injection qua tham số tìm kiếm khóa học chưa sanitize","KhoaHocController.cs:89","1h","In Progress","xuanbac9968","SonarQube"),
    ("NNL-003","Bug","MAJOR","Medium","Race condition Duyet() – 2 admin duyệt cùng lúc, trạng thái không nhất quán","ThanhToanController.cs:156","2h","In Progress","xuanbac9968","SonarQube"),
    ("NNL-004","Vulnerability","CRITICAL","High","Missing CSRF (ValidateAntiForgeryToken) trên POST /ThanhToan/Duyet","ThanhToanController.cs:143","15m","Open","xuanbac9968","SonarQube"),
    ("NNL-005","Vulnerability","MAJOR","High","Hard-coded credentials Admin@123/Gv@123/Hv@123 trong SeedData.cs","SeedData.cs:23","1h","In Progress","xuanbac9968","SonarQube"),
    ("NNL-006","Bug","MAJOR","Medium","HttpClient tạo mới mỗi request ChatService – resource leak, socket exhaustion","ChatService.cs:34","2h","Open","xuanbac9968","SonarQube"),
    ("NNL-007","Vulnerability","MAJOR","Medium","API token Grok lộ trong appsettings.json nếu commit public repo","appsettings.json:8","30m","Done","xuanbac9968","SonarQube"),
    ("NNL-008","Bug","MINOR","Low","DiemTongKet không validate đầu vào – kết quả có thể ngoài [0,10]","Services/DiemService.cs:67","30m","Open","xuanbac9968","SonarQube"),
    ("NNL-009","Vulnerability","MAJOR","Medium","Missing Content-Security-Policy header – nguy cơ XSS","Program.cs:45","1h","Open","xuanbac9968","SonarQube + ZAP"),
    ("NNL-010","Bug","MAJOR","Medium","LichHocHelper.ExpandToSessions() không validate NgayKetThuc < NgayBatDau","Helpers/LichHocHelper.cs:28","30m","In Progress","xuanbac9968","SonarQube"),
    ("NNL-011","Vulnerability","CRITICAL","High","File upload chỉ check extension – bypass bằng rename .exe → .jpg","ThanhToanController.cs:198","2h","Open","xuanbac9968","SonarQube"),
    ("NNL-012","Vulnerability","MAJOR","Medium","Session fixation – không regenerate session ID sau đăng nhập","Program.cs:67","1h","Open","xuanbac9968","SonarQube"),
    ("NNL-013","Bug","MINOR","Low","PhanCong conflict không cover IsActive=false rồi bật lại","AdminController.cs:312","30m","Done","xuanbac9968","SonarQube"),
    ("NNL-014","Vulnerability","MAJOR","Medium","Missing X-Frame-Options / CSP frame-ancestors – clickjacking risk","Program.cs:45","15m","Open","xuanbac9968","SonarQube + ZAP"),
    ("NNL-015","Bug","MAJOR","Medium","GoiYKhoaHocService HttpClient thiếu Timeout – request treo indefinitely","Services/GoiYKhoaHocService.cs:55","1h","In Progress","xuanbac9968","SonarQube"),
    ("NNL-016","Vulnerability","MINOR","Low","Missing X-Content-Type-Options: nosniff header","Program.cs:45","10m","Open","xuanbac9968","OWASP ZAP"),
    ("NNL-017","Bug","MINOR","Low","ExcelService workbook không dispose đúng khi SaveAs() ném exception","Services/ExcelService.cs:88","30m","Open","xuanbac9968","SonarQube"),
    ("NNL-018","Vulnerability","MAJOR","Medium","CDN scripts thiếu SRI (Subresource Integrity) – tấn công nếu CDN bị compromise","Views/Shared/_Layout.cshtml:12","2h","Open","xuanbac9968","OWASP ZAP"),
    ("NNL-019","Bug","MINOR","Low","Phân trang Index() không reset trang 1 khi filter thay đổi","KhoaHocController.cs:124","30m","Open","xuanbac9968","SonarQube"),
    ("NNL-020","Bug","MINOR","Low","BuoiHocViewModel.TenThu sai Chủ nhật – off-by-one DayOfWeek vs ThuTrongTuan","Helpers/LichHocHelper.cs:45","30m","Open","xuanbac9968","SonarQube"),
]

sev_fill = {"CRITICAL":C["risk_critical"],"MAJOR":C["risk_high"],"MINOR":C["risk_low"]}
type_fill = {"Bug":PatternFill("solid",fgColor="FFCDD2"),"Vulnerability":PatternFill("solid",fgColor="FFF9C4")}
status_f2 = {"Open":C["fail_red"],"In Progress":C["warn_yellow"],"Done":C["pass_green"]}

for i, (jid, typ, sev, pri, desc, comp, effort, status, asgn, src) in enumerate(BUGS):
    row = 5 + i
    base = C["row_light"] if i%2==0 else C["row_alt"]
    tfill = type_fill.get(typ, base)
    sfill = sev_fill.get(sev, base)
    stfill = status_f2.get(status, base)

    set_row(ws3, row,
            [jid, typ, sev, pri, desc, comp, effort, status, asgn, src],
            fills=[base, tfill, sfill, base, base, base, base, stfill, base, base],
            fonts=[BOLD10, BOLD10, BOLD10, NORMAL, NORMAL, SMALL, SMALL, BOLD10, NORMAL, SMALL],
            aligns=[CENTER, CENTER, CENTER, CENTER, LEFT, LEFT, CENTER, CENTER, CENTER, LEFT],
            height=36)

ws3.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
#  SHEET 4: KẾT QUẢ KIỂM THỬ TỔNG HỢP
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Kết Quả Tổng Hợp")

title_row(ws4, 1, "✅ KẾT QUẢ KIỂM THỬ TỔNG HỢP – QuanLyKhoaHoc5", 7, C["header_purple"])

col_widths(ws4, [22, 18, 18, 18, 18, 18, 20])
ws4.row_dimensions[2].height = 8

# Metrics table
headers4 = ["Loại kiểm thử", "Công cụ", "Tổng TH", "Passed", "Failed", "Pass Rate", "Trạng thái"]
set_row(ws4, 3, headers4, C["header_purple"], WHITE, CENTER, 36)

RESULTS = [
    ("Unit Tests (xUnit)","xUnit + Moq + coverlet",531,531,0,"100%","✅ PASSED"),
    ("Code Coverage","coverlet + SonarQube","–","82.9%","–","≥80% ✅","✅ PASSED (QG)"),
    ("E2E Tests (Katalon)","Katalon Studio 11.1.3.0",6,6,0,"100%","✅ PASSED"),
    ("Load Test 100u (JMeter)","JMeter 5.6.3",2000,2000,0,"100%","✅ OK (P95:631ms)"),
    ("Load Test 200u (JMeter)","JMeter 5.6.3",2000,2000,0,"100%","✅ OK (P95:1027ms)"),
    ("Load Test 300u (JMeter)","JMeter 5.6.3",2000,2000,0,"100%","✅ OK (P95:769ms)"),
    ("Security Scan (ZAP)","OWASP ZAP 2.16.0","–","0 High","3 Med","No High ✅","⚠️ 3 Medium"),
    ("Code Quality (Sonar)","SonarQube 26.5.0","–","QG PASSED","–","Coverage 82.9%","✅ Quality Gate OK"),
    ("CI/CD Pipeline","GitHub Actions","5 Jobs","5 Jobs","0","100%","✅ ALL PASSED"),
]

for i, row_data in enumerate(RESULTS):
    row = 4 + i
    base = C["row_light"] if i%2==0 else C["row_alt"]
    sfill = C["pass_green"] if "PASSED" in str(row_data[6]) or "OK" in str(row_data[6]) else C["warn_yellow"]
    fills = [base, base, base, base, base, base, sfill]
    set_row(ws4, row, list(row_data), fills=fills,
            fonts=[BOLD10]+[NORMAL]*5+[BOLD10],
            aligns=[LEFT, LEFT, CENTER, CENTER, CENTER, CENTER, CENTER], height=28)

ws4.row_dimensions[4+len(RESULTS)].height = 12

# Summary conclusion
cr = 4 + len(RESULTS) + 1
ws4.merge_cells(f"A{cr}:G{cr}")
c = ws4.cell(row=cr, column=1, value="🏆 KẾT LUẬN TỔNG QUÁT")
c.fill = C["header_dark"]; c.font = TITLE_F; c.alignment = CENTER
ws4.row_dimensions[cr].height = 30

cr += 1
conclusion = [
    "✅ Unit Tests: 531 tests PASSED (100%), coverage 82.9% – Quality Gate PASSED",
    "✅ E2E Tests: 6/6 test cases PASSED – Katalon Studio 11.1.3.0",
    "✅ Performance: 0 errors ở 100/200/300 users, P95 < SLA 2000ms",
    "⚠️ Security: 0 High, 3 Medium (CSP/Clickjacking/SRI) – cần fix trước production",
    "⚠️ Bug Tracking: 14 issues Open (3 Critical cần xử lý ngay: NNL-004, NNL-009, NNL-011)",
    "✅ CI/CD: GitHub Actions pipeline 5 jobs – tự động hóa toàn bộ quy trình kiểm thử",
]
conc_fills = [C["pass_green"], C["pass_green"], C["pass_green"], C["warn_yellow"], C["warn_yellow"], C["pass_green"]]
for j, txt in enumerate(conclusion):
    ws4.merge_cells(f"A{cr+j}:G{cr+j}")
    c = ws4.cell(row=cr+j, column=1, value=txt)
    c.fill = conc_fills[j]; c.font = Font(size=10, bold=("✅" in txt))
    c.alignment = LEFT; c.border = thin_border()
    ws4.row_dimensions[cr+j].height = 22

ws4.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
#  SHEET 5: BIỂU ĐỒ RISK MATRIX
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Risk Matrix")

title_row(ws5, 1, "📊 RISK MATRIX – Ma trận rủi ro 5×5", 7, C["header_orange"])
col_widths(ws5, [14, 14, 14, 14, 14, 14, 14])
ws5.row_dimensions[2].height = 8

# Matrix header
set_row(ws5, 3, ["Xác suất \\ Tác động", "1 – Không đáng kể", "2 – Nhẹ", "3 – Trung bình", "4 – Nghiêm trọng", "5 – Thảm họa", "Màu tham chiếu"],
        C["header_blue"], WHITE, CENTER, 36)

MATRIX = [
    ("5 – Gần chắc", [5,10,15,20,25]),
    ("4 – Khả năng cao", [4,8,12,16,20]),
    ("3 – Có thể", [3,6,9,12,15]),
    ("2 – Khó xảy ra", [2,4,6,8,10]),
    ("1 – Hiếm khi", [1,2,3,4,5]),
]

cell_fill = {
    range(1,5):  PatternFill("solid", fgColor="C8E6C9"),
    range(5,9):  PatternFill("solid", fgColor="FFF9C4"),
    range(9,13): PatternFill("solid", fgColor="FFE0B2"),
    range(13,21):PatternFill("solid", fgColor="FFCDD2"),
    range(21,30):PatternFill("solid", fgColor="EF9A9A"),
}
def get_cell_fill(val):
    for r, f in cell_fill.items():
        if val in r: return f
    return PatternFill("solid", fgColor="EF9A9A")

for i, (label, vals) in enumerate(MATRIX):
    row = 4 + i
    ws5.row_dimensions[row].height = 28
    c = ws5.cell(row=row, column=1, value=label)
    c.fill = C["header_blue"]; c.font = WHITE; c.alignment = CENTER; c.border = thin_border()
    for j, v in enumerate(vals):
        cell = ws5.cell(row=row, column=2+j, value=v)
        cell.fill = get_cell_fill(v)
        cell.font = Font(bold=True, size=12)
        cell.alignment = CENTER; cell.border = thin_border()
    # Color legend
    c = ws5.cell(row=row, column=7, value=["🟢 Thấp (1-4)","🟡 Trung bình (5-8)","🟠 Cao (9-12)","🔴 Nghiêm trọng (13-20)","🔴 Cực kỳ nguy hiểm (21-25)"][i])
    c.font = Font(size=10, bold=True); c.alignment = LEFT; c.border = thin_border()

# Plot actual risks on matrix
ws5.row_dimensions[10].height = 12
ws5.merge_cells("A11:G11")
c = ws5["A11"]; c.value = "📍 VỊ TRÍ CÁC RỦI RO TRÊN MATRIX"
c.fill = C["header_dark"]; c.font = TITLE_F; c.alignment = CENTER; ws5.row_dimensions[11].height = 28

set_row(ws5, 12, ["ID Rủi ro", "Chức năng", "XS", "MĐAH", "Điểm", "Mức độ", "Jira Issues"],
        C["header_blue"], WHITE, CENTER, 32)

RISK_SUMMARY = [(rid, mod, xs, md) for rid, mod, _, xs, md, *rest in RISKS]
for i, (rid, mod, xs, md) in enumerate(RISK_SUMMARY):
    score = xs * md
    _, lbl, rfill, icon = get_risk(xs, md)
    row = 13 + i
    jira = "NNL-" + str(i+1).zfill(3)
    set_row(ws5, row,
            [rid, mod, xs, md, score, f"{icon} {lbl}", jira],
            fills=[C["row_light"] if i%2==0 else C["row_alt"]]*5+[rfill]+[C["row_alt"]],
            aligns=[CENTER]*4+[CENTER, CENTER, LEFT], height=22)

ws5.freeze_panes = "A4"

# Save
wb.save(OUTPUT)
print(f"\n✅ File Excel đã được tạo: {OUTPUT}")
print(f"   📄 5 sheets:")
print(f"   1. Phân Tích Rủi Ro – {len(RISKS)} rủi ro với ma trận 5×5")
print(f"   2. Chiến Lược Kiểm Thử – {len(STRATEGIES)} chiến lược theo module")
print(f"   3. Bug Tracker (Jira) – {len(BUGS)} issues từ SonarQube + ZAP")
print(f"   4. Kết Quả Tổng Hợp – {len(RESULTS)} loại kiểm thử")
print(f"   5. Risk Matrix – Ma trận rủi ro 5×5 + vị trí các rủi ro")
